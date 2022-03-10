import json
import HVAC_CalcEngine_DictBased as HVAC
from copy import copy, deepcopy
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import os
import datetime
import pdb

#%% old stuff
# def generate_variation(value, inputs_dict, mitigation, results, variations):
    
#     hvac_obj = HVAC.HVAC_Internal_Calculation(inputs_dict)
#     summary_values = hvac_obj.get_summary_values()
    
#     for result in results:
#         variations[mitigation]['absolute'][result].append(summary_values['total'][result+'_total'])
        
#         try:
#             percentage = summary_values['total'][result+'_total']/variations[mitigation]['absolute'][result][0]
#         except:
#             percentage = 0
#         variations[mitigation]['percentage'][result].append(percentage)
    
#     total  = (summary_values['total']['fans_total']+summary_values['total']['preheat_total']+
#               summary_values['total']['heating_total']+summary_values['total']['cooling_total']+
#               summary_values['total']['dehum_total']+summary_values['total']['hum_total'])
#     variations[mitigation]['absolute']['total'].append(total)
    
#     percentage = total/variations[mitigation]['absolute']['total'][0]
#     variations[mitigation]['percentage']['total'].append(percentage)
    
#     variations[mitigation]['value'].append(value)
#     pdb.set_trace()
#     return variations



# def supply_airflow_variations(airflows, inputs_dict, variations, mitigation, results):

#     for airflow in airflows:
#         inputs_dict_var = deepcopy(inputs_dict)
#         inputs_dict_var['model_inputs']['General']['Supply Airflow'] = inputs_dict_var['model_inputs']['General']['Supply Airflow']*airflow
        
#         variations = generate_variation(airflow, inputs_dict_var, mitigation, results, variations)
        
#     return variations
    
# def temperature_controlband_variations(controlbands, inputs_dict, variations, mitigation, results):
#     for controlband in controlbands:
#         inputs_dict_var = deepcopy(inputs_dict)
#         for i in range(1,4):
#             #if no setpoint is given for no.2 and 3
#             if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Lower Band']) == 'nan':
#                 inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Lower Band'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Temperature Lower Band']
#             if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Upper Band']) == 'nan':
#                 inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Upper Band'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Temperature Upper Band']
            
#             inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Lower Band'] += -controlband
#             inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Upper Band'] += controlband
        
#         variations = generate_variation(controlband, inputs_dict_var, mitigation, results, variations)
        
#     return variations

# def humidity_controlband_variations(controlbands, inputs_dict, variations, mitigation, results):
#     for controlband in controlbands:
#         inputs_dict_var = deepcopy(inputs_dict)
#         for i in range(1,4):
#             #if no setpoint is given for no.2 and 3
#             if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Humidity Lower Band']) == 'nan':
#                 inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Humidity Lower Band'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Humidity Lower Band']
#             if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Humidity Upper Band']) == 'nan':
#                 inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Humidity Upper Band'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Humidity Upper Band']
            
#             inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Humidity Lower Band'] += -controlband
#             inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Humidity Upper Band'] += controlband
        
#         variations = generate_variation(controlband, inputs_dict_var, mitigation, results, variations)
        
#     return variations
    
# def temperature_setpoint_variations(setpoints, inputs_dict, variations, mitigation, results):
#     for setpoint in setpoints:
#         inputs_dict_var = deepcopy(inputs_dict)
#         for i in range(1,4):
#             #if no setpoint is given for no.2 and 3
#             if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Lower Band']) == 'nan':
#                 inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Lower Band'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Temperature Lower Band']
#             if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Upper Band']) == 'nan':
#                 inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Upper Band'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Temperature Upper Band']
#             if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Setpoint']) == 'nan':
#                 inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Setpoint'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Temperature Setpoint']
                    
#             inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Lower Band'] += setpoint
#             inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Upper Band'] += setpoint
#             inputs_dict_var['model_inputs']['Set Point Profile '+str(i)]['Room Temperature Setpoint']   += setpoint
        
#         variations = generate_variation(setpoint, inputs_dict_var, mitigation, results, variations)
           
#     return variations

# def create_variations(summary_values, variant):
    
#     global mitigations
#     mitigations = ['supply airflow', 'temperature control bands', 'humidity control bands', 'temperature setpoints']
#     results     = ['fans', 'preheat', 'heating', 'cooling', 'dehum', 'hum']
#     total       = (summary_values['total']['fans_total']+summary_values['total']['preheat_total']+
#                    summary_values['total']['heating_total']+summary_values['total']['cooling_total']+
#                    summary_values['total']['dehum_total']+summary_values['total']['hum_total'])
#     variations = {}
#     for mitigation in mitigations:
#         variations[mitigation] = {}
#         variations[mitigation]['absolute'] = {'fans':    [summary_values['total']['fans_total']],
#                                               'preheat': [summary_values['total']['preheat_total']],
#                                               'heating': [summary_values['total']['heating_total']],
#                                               'cooling': [summary_values['total']['cooling_total']],
#                                               'dehum':   [summary_values['total']['dehum_total']],
#                                               'hum':     [summary_values['total']['hum_total']],
#                                               'total':   [total]
#                                               }
#         variations[mitigation]['percentage'] = {'fans':  [1],
#                                               'preheat': [1],
#                                               'heating': [1],
#                                               'cooling': [1],
#                                               'dehum':   [1],
#                                               'hum':     [1],
#                                               'total':   [1]
#                                               }
#         if mitigation==mitigations[0]:
#             variations[mitigation]['value'] = [1] 
#         if mitigation==mitigations[1] or mitigation==mitigations[2] or mitigation==mitigations[3]:
#             variations[mitigation]['value'] = [0] 

#     airflows                 = [0.95, 0.9, 0.85, 0.8, 0.75, 0.7]
#     temperature_controlbands = [0.5, 1, 1.5, 2, 10]
#     humidity_controlbands    = [0.05, 0.1, 0.15, 0.2]
#     temperature_setpoints    = [-1, 1]
    
#     variations = supply_airflow_variations(airflows, variant, variations, mitigations[0], results)
#     variations = temperature_controlband_variations(temperature_controlbands, variant, variations, mitigations[1], results)
#     variations = humidity_controlband_variations(humidity_controlbands, variant, variations, mitigations[2], results)
#     variations = temperature_setpoint_variations(temperature_setpoints, variant, variations, mitigations[3], results)
#     return variations
#%%
def generate_variation(inputs_dict, measure, result_names, variations):
    
    hvac_obj = HVAC.HVAC_Internal_Calculation(inputs_dict)
    summary_values = hvac_obj.get_summary_values()
    
    for result in result_names:
        variations['absolute'][result].append(summary_values['total'][result+'_total'])
        
        try:
            percentage = summary_values['total'][result+'_total']/variations['absolute'][result][0]
        except:
            percentage = 0
        variations['percentage'][result].append(percentage)
    
    total  = (summary_values['total']['fans_total']+summary_values['total']['preheat_total']+
              summary_values['total']['heating_total']+summary_values['total']['cooling_total']+
              summary_values['total']['dehum_total']+summary_values['total']['hum_total'])
    variations['absolute']['total'].append(total)
    
    percentage = total/variations['absolute']['total'][0]
    variations['percentage']['total'].append(percentage)
    
    variations['measure'].append(measure)
    # pdb.set_trace()
    return variations


#%%
def create_variation(inputs_dict_base, measures_dict):
    inputs_dict_var = deepcopy(inputs_dict_base)
    #%% widen temperature band
    if measures_dict['widen_temp_band']:
        for iSetpoint in range(1,4):
            #if no setpoint is given for no.2 and 3
            if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Lower Band']) == 'nan':
                inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Lower Band'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Temperature Lower Band']
            if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Upper Band']) == 'nan':
                inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Upper Band'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Temperature Upper Band']
            
            inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Lower Band'] += -1
            inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Upper Band'] += 1
    
    #%% reduce setpoint
    if measures_dict['reduce_temp_setpoint']:
        for iSetpoint in range(1,4):
            #if no setpoint is given for no.2 and 3
            if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Lower Band']) == 'nan':
                inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Lower Band'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Temperature Lower Band']
            if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Upper Band']) == 'nan':
                inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Upper Band'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Temperature Upper Band']
            if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Setpoint']) == 'nan':
                inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Setpoint'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Temperature Setpoint']
                
            inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Lower Band'] -= 1
            inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Upper Band'] -= 1
            inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Setpoint']   -= 1
    
    #%% increase setpoint
    if measures_dict['increase_temp_setpoint']:
        for iSetpoint in range(1,4):
            #if no setpoint is given for no.2 and 3
            if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Lower Band']) == 'nan':
                inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Lower Band'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Temperature Lower Band']
            if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Upper Band']) == 'nan':
                inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Upper Band'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Temperature Upper Band']
            if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Setpoint']) == 'nan':
                inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Setpoint'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Temperature Setpoint']
                
            inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Lower Band'] += 1
            inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Upper Band'] += 1
            inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Temperature Setpoint']   += 1
            
    #%% widen humidity band
    if measures_dict['widen_humidity_band']:
        for iSetpoint in range(1,4):
            #if no setpoint is given for no.2 and 3
            if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Humidity Lower Band']) == 'nan':
                inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Humidity Lower Band'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Humidity Lower Band']
            if str(inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Humidity Upper Band']) == 'nan':
                inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Humidity Upper Band'] = inputs_dict_var['model_inputs']['Set Point Profile 1']['Room Humidity Upper Band']
            
            inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Humidity Lower Band'] += -0.05
            inputs_dict_var['model_inputs']['Set Point Profile '+str(iSetpoint)]['Room Humidity Upper Band'] += 0.05
        
    #%% reduce air flow
    if measures_dict['reduce_air_flow']:
        inputs_dict_var['model_inputs']['General']['Supply Airflow'] = inputs_dict_var['model_inputs']['General']['Supply Airflow']*0.7
        inputs_dict_var['model_inputs']['General']['Return Airflow'] = inputs_dict_var['model_inputs']['General']['Return Airflow']*0.7
    
    #%% incorporate run around coild heat recovery
    if measures_dict['incorporate_run_around_coil_heat_recovery']:
        if inputs_dict_var['model_inputs']['General']['HVAC Module 1'] == 'Mixing Box' or inputs_dict_var['model_inputs']['General']['HVAC Module 1'] == 'Thermal Wheel' or inputs_dict_var['model_inputs']['General']['HVAC Module 1'] == 'Plate Heat Exchanger':
            inputs_dict_var['model_inputs']['General']['HVAC Module 1'] == 'Run Around Coil'
        else:
            for iHVAC in range(2,10):
                inputs_dict_var['model_inputs']['General']['HVAC Module ' + str(iHVAC)] = inputs_dict_var['model_inputs']['General']['HVAC Module ' + str(iHVAC-1)]
            inputs_dict_var['model_inputs']['General']['HVAC Module 1'] = 'Run Around Coil'
    
    #%% incorporate mixing box heat recovery
    if measures_dict['incorporate_mixing_box_recovery']:
        if inputs_dict_var['model_inputs']['General']['HVAC Module 1'] == 'Run Around Coil' or inputs_dict_var['model_inputs']['General']['HVAC Module 1'] == 'Thermal Wheel' or inputs_dict_var['model_inputs']['General']['HVAC Module 1'] == 'Plate Heat Exchanger':
            inputs_dict_var['model_inputs']['General']['HVAC Module 1'] == 'Mixing Box'
        else:
            for iHVAC in range(2,10):
                inputs_dict_var['model_inputs']['General']['HVAC Module ' + str(iHVAC)] = inputs_dict_var['model_inputs']['General']['HVAC Module ' + str(iHVAC-1)]
            inputs_dict_var['model_inputs']['General']['HVAC Module 1'] = 'Mixing Box'
            
    #%% incorporate different time-schedule
    if measures_dict['tweak_schedule_1']:
        time_1     = 22
        time_2     = 5
        percentage = 0.8
        
        diff       = 24-time_1 + time_2
        hours      = []
        for j in range(365):
            hours.append(time_1+(j*24))
            for k in range(diff):
                hours.append(hours[-1]+1)
        new_schedule = deepcopy(inputs_dict_var['run_schedule']['Run_Schedule_Profile'])
        for i in range(24):
            if i+1<=time_2:
                new_schedule[i] = percentage
        for hour in hours:
            try:
                new_schedule[hour-1] = percentage 
            except:
                pass
        inputs_dict_var['run_schedule']['Run_Schedule_Profile'] = new_schedule
        
    #%% incorporate different time-schedule
    if measures_dict['tweak_schedule_2']:
        time_1     = 22
        time_2     = 5
        percentage = 0.6
        
        diff       = 24-time_1 + time_2
        hours      = []
        for j in range(365):
            hours.append(time_1+(j*24))
            for k in range(diff):
                hours.append(hours[-1]+1)
        new_schedule = deepcopy(inputs_dict_var['run_schedule']['Run_Schedule_Profile'])
        for i in range(24):
            if i+1<=time_2:
                new_schedule[i] = percentage
        for hour in hours:
            try:
                new_schedule[hour-1] = percentage 
            except:
                pass
        inputs_dict_var['run_schedule']['Run_Schedule_Profile'] = new_schedule
        
    #%% incorporate different time-schedule
    if measures_dict['tweak_schedule_3']:
        time_1     = 22
        time_2     = 5
        percentage = 0.4
        
        diff       = 24-time_1 + time_2
        hours      = []
        for j in range(365):
            hours.append(time_1+(j*24))
            for k in range(diff):
                hours.append(hours[-1]+1)
        new_schedule = deepcopy(inputs_dict_var['run_schedule']['Run_Schedule_Profile'])
        for i in range(24):
            if i+1<=time_2:
                new_schedule[i] = percentage
        for hour in hours:
            try:
                new_schedule[hour-1] = percentage 
            except:
                pass
        inputs_dict_var['run_schedule']['Run_Schedule_Profile'] = new_schedule
    
    return inputs_dict_var
  
  
#%%
def read_measure_combination(measure_combination):
    measures = {}
    measures['string'] = measure_combination
    if not measure_combination.find('A1') == -1:
        measures['widen_temp_band'] = True
    else:
        measures['widen_temp_band'] = False
        
    if not measure_combination.find('A2') == -1:
        measures['reduce_temp_setpoint'] = True
    else:
        measures['reduce_temp_setpoint'] = False
        
    if not measure_combination.find('A3') == -1:
        measures['increase_temp_setpoint'] = True
    else:
        measures['increase_temp_setpoint'] = False
    
    if not measure_combination.find('B1') == -1:
        measures['widen_humidity_band'] = True
    else:
        measures['widen_humidity_band'] = False
        
    if not measure_combination.find('C1') == -1:
        measures['reduce_air_flow'] = True
    else:
        measures['reduce_air_flow'] = False
        
    if not measure_combination.find('D1') == -1:
        measures['incorporate_run_around_coil_heat_recovery'] = True
    else:
        measures['incorporate_run_around_coil_heat_recovery'] = False
        
    if not measure_combination.find('D2') == -1:
        measures['incorporate_mixing_box_recovery'] = True
    else:
        measures['incorporate_mixing_box_recovery'] = False
        
    if not measure_combination.find('E1') == -1:
        measures['tweak_schedule_1'] = True
    else:
        measures['tweak_schedule_1'] = False
        
    if not measure_combination.find('E2') == -1:
        measures['tweak_schedule_2'] = True
    else:
        measures['tweak_schedule_2'] = False
        
    if not measure_combination.find('E3') == -1:
        measures['tweak_schedule_3'] = True
    else:
        measures['tweak_schedule_3'] = False
    
    return measures

    
def create_variations_batch(summary_values, inputs_dict_var_base):
    possible_measures =    ['C1',
                            'B1C1',
                            'A1C1',
                            'A1B1C1',
                            'A1C1A2',
                            'A1C1A3',
                            'B1C1D2',
                            'A1C1D2',
                            'A1B1C1D2',
                            'A2',
                            'A1A2',
                            'D2',
                            'A1',
                            'A1B1D2',
                            'B1',
                            'A3',
                            'A1B1',
                            'A1A3',
                            'A1C1D1',
                            'A1B1C1D1',
                            'B1C1D1',
                            'A1B1D1',
                            'D1',
                            'E1',
                            'E2',
                            'E3']
                            
    result_names     = ['fans', 'preheat', 'heating', 'cooling', 'dehum', 'hum']
    total       = (summary_values['total']['fans_total']+summary_values['total']['preheat_total']+
                   summary_values['total']['heating_total']+summary_values['total']['cooling_total']+
                   summary_values['total']['dehum_total']+summary_values['total']['hum_total'])
    
    # check if Run Around Coil or Mixing Box exists and if so all possible measures with D1 or D2 are skipped
    D1 = True
    D2 = True
    if inputs_dict_var_base['model_inputs']['General']['HVAC Module 1'] == 'Run Around Coil':
        D1 = False
    if inputs_dict_var_base['model_inputs']['General']['HVAC Module 1'] == 'Mixing Box':
        D2 = False
    
    # initialise variation results dictionary
    variations = {}
    variations['absolute'] = {'fans':    [summary_values['total']['fans_total']],
                              'preheat': [summary_values['total']['preheat_total']],
                              'heating': [summary_values['total']['heating_total']],
                              'cooling': [summary_values['total']['cooling_total']],
                              'dehum':   [summary_values['total']['dehum_total']],
                              'hum':     [summary_values['total']['hum_total']],
                              'total':   [total]
                              }
    variations['percentage'] = {'fans':    [1],
                                'preheat': [1],
                                'heating': [1],
                                'cooling': [1],
                                'dehum':   [1],
                                'hum':     [1],
                                'total':   [1]
                                }
    variations['measure'] = ['-']
    variations['input_dicts_var'] = [inputs_dict_var_base]
    variations['measure_bool'] = [{'widen_temp_band'                            : False,
                                  'reduce_temp_setpoint'                        : False,
                                  'increase_temp_setpoint'                      : False,
                                  'widen_humidity_band'                         : False,
                                  'reduce_air_flow'                             : False,
                                  'incorporate_run_around_coil_heat_recovery'   : False,
                                  'incorporate_mixing_box_recovery'             : False,
                                  'tweak_schedule_1'                            : False,
                                  'tweak_schedule_2'                            : False,
                                  'tweak_schedule_3'                            : False,
                                  'string'                                      : '-'}]
    # if mitigation==mitigations[0]:
    #     variations[mitigation]['value'] = [1] 
    # if mitigation==mitigations[1] or mitigation==mitigations[2] or mitigation==mitigations[3]:
    #     variations[mitigation]['value'] = [0] 


    for measure in possible_measures:
        #            ~D1              D1 found
        if not (D1 == False and not measure.find('D1') == -1) and not (D2 == False and not measure.find('D2') == -1):
            print(measure)
            measures_dict = read_measure_combination(measure)
            variations['measure_bool'].append(measures_dict)
            inputs_dict_var_measure = create_variation(inputs_dict_var_base, measures_dict)
            variations['input_dicts_var'].append(inputs_dict_var_measure)
            variations = generate_variation(inputs_dict_var_measure, measure, result_names, variations)
            
            # pdb.set_trace()
        else:
            print(measure + ' skipped!')
            
    return variations

def write_summary_values(model_file_path, summary_results_sheet, summary_values):

    wb = load_workbook(model_file_path)
    wsh = wb[summary_results_sheet]

    #writing min values
    min_keys = ['fans_min', 'preheat_min', 'dehum_min', 'hum_min', 'heating_min', 'cooling_min']
    for i in range(6):
        wsh.cell(column = i+3, row = 7, value = summary_values['min'][min_keys[i]]) 

    #writing max values
    max_keys = ['fans_max', 'preheat_max', 'dehum_max', 'hum_max', 'heating_max', 'cooling_max']
    for i in range(6):
        wsh.cell(column = i+3, row = 8, value = summary_values['max'][max_keys[i]]) 

    #writing avg values
    avg_keys = ['fans_avg', 'preheat_avg', 'dehum_avg', 'hum_avg', 'heating_avg', 'cooling_avg']
    for i in range(6):
        wsh.cell(column = i+3, row = 9, value = summary_values['avg'][avg_keys[i]]) 

    #writing total values
    total_keys = ['fans_total', 'preheat_total', 'dehum_total', 'hum_total', 'heating_total', 'cooling_total']
    for i in range(6):
        wsh.cell(column = i+3, row = 10, value = summary_values['total'][total_keys[i]]) 

    #writing kpis
    kpi_keys = ['hvac_total_energy_area', 'hvac_heating', 'hvac_cooling', 'hvac_humidification', 'hvac_dehumidification', 'hvac_fans']
    for i in range(6):
        wsh.cell(column = 10, row = i+13, value = summary_values['kpi'][kpi_keys[i]]) 

    #writing hours within
    within_keys = ['tdb_hours','rh_hours','tdb_rh_hours']
    for i in range(3):
        wsh.cell(column = i+2, row = 15, value = summary_values['within'][within_keys[i]])

    #writing hours outside
    outside_keys = ['tdb_hours','rh_hours','tdb_rh_hours']
    for i in range(3):
        wsh.cell(column = i+2, row = 18, value = summary_values['outside'][outside_keys[i]])

    # Save the file
    wb.save(model_file_path)

def write_detailed_result(model_file_path, detailed_results_sheet, detailed_result):

    #load workbook and point to Model Results Detailed sheet
    wb = load_workbook(model_file_path)
    wsh = wb[detailed_results_sheet]

    # clear the existing values
    for row in wsh['B2:AR8990']:
        for cell in row:
            cell.value = None

    #write result
    output_cols = []
    for i in range(2,100):
        if wsh.cell(1, i).value is None:
            break
        output_cols.append(wsh.cell(1, i).value)

    #output_cols = ['Room_Sensible_Load[kW]', 'Room_Latent_Load[kW]', 'Room_Ratio_Line', 'Supply_Fan_Power[kW]', 'Return_Fan_Power[kW]', 'Airflow_Fresh[kg/s]', 'Airflow_Return[kg/s]', 'Airflow_Total_Supply[kg/s]', 'Preheat_Load[kW_S]', 'Heating_Load[kW_S]', 'Reheat_Load[kW_S]', 'Cooling_Load[kW_L]', 'Cooling_Load[kW_S]', 'Dehumidification_Load[kW_L]', 'Dehumidification_Load[kW_S]', 'Humidification_Load[kW_L]', 'Humidification_Load[kW_S]', 'Humidification_Water[L]', 'Fresh_Air_Tdb', 'Fresh_Air_RH', 'Return_Air_Tdb', 'Return_Air_RH', 'Post_Preheat_Tdb', 'Post_Preheat_RH', 'Post_HR_Air_Tdb', 'Post_HR_Air_RH', 'Post_Heat_Tdb', 'Post_Heat_RH', 'Post_Cool_Tdb', 'Post_Cool_RH', 'Post_Dehum_Tdb', 'Post_Dehum_RH', 'Post_Reheat_Tdb', 'Post_Reheat_RH', 'Post_Hum_Tdb', 'Post_Hum_RH', 'Required_Supply_Air_Tdb', 'Required_Supply_Air_RH', 'Final_Supply_Air_Tdb', 'Final_Supply_Air_RH', 'Final_Room_Air_Tdb', 'Final_Room_Air_RH', 'AHU_Mode']
    for i, key in enumerate(output_cols):
        if len(detailed_result[key]) > 10: # If the result has values
                for j in range (0,8760,1):
                    wsh.cell(column = i+2, row = j+2, value = detailed_result[key][j]) 
                    #if j==4: break

    # Save the file
    wb.save(model_file_path)

def write_variations_result(model_file_path, variations_result_sheet, variation_result, basecase_result):
    
    #read the file
    wb = load_workbook(model_file_path)
    wsh = wb[variations_result_sheet]
    
    #keys correponsding to columns in the variation results sheet
    summary_keys = ['fans_total', 'preheat_total', 'dehum_total', 'hum_total', 'heating_total', 'cooling_total']
    
    #Update the variations model inputs
    #for each key
    for i in range(7, 50):
        
        #if all the variations are populated, break
        if wsh.cell(i, 2).value is None:
            break

        key = wsh.cell(i, 2).value
        
        #for each column
        for j in range(5, 11):
            
            if key == 'Base Scenario':
                #write base case results
                wsh.cell(column = j, row = i, value = basecase_result[summary_keys[j-5]]) 
            else:   
                #write variations results
                wsh.cell(column = j, row = i, value = variation_result[key][summary_keys[j-5]]) 

    # Save the file
    wb.save(model_file_path)

def write_summary_batch(inputs_dicts, summary_values_cases, model_file_path):
    wb = load_workbook(model_file_path)
    wsh = wb['Transpose']
    
    n_AHUs = len(inputs_dicts)
    
    # delete rows that are unnessecary
    for iRow in range(3,10000):
        for iCol in range(1,77):
            wsh.cell(column = iCol, row = iRow, value = '')
    # wsh.delete_rows(3,10000)
    
    # copy the first row as many times as there are AHUs
    for iRow in range(1,n_AHUs):
        for iCol in range(1,77):
            wsh.cell(column = iCol, row = iRow+2, value = wsh.cell(column = iCol, row = 2).value)
            # wsh.cell(column = iCol, row = iRow+2).data_type = wsh.cell(column = iCol, row = 2).data_type
            # wsh.cell(column = iCol, row = iRow+2).encoding = wsh.cell(column = iCol, row = 2).encoding
            
    
    # now fill the values
    for iAHU in range(n_AHUs):
        iCol = 1
        # Runtime
        wsh.cell( column = iCol, row = iAHU+2, value = str(datetime.datetime.now().strftime('%y-%m-%d %H-%M')) )
        iCol += 1
        
        # Version
        wsh.cell( column = iCol, row = iAHU+2, value = 0 )
        iCol += 1
        
        # Climate1
        wsh.cell( column = iCol, row = iAHU+2, value = False )
        iCol += 1
        
        # Climate2
        wsh.cell( column = iCol, row = iAHU+2, value = False )
        iCol += 1
        
        # Site
        wsh.cell( column = iCol, row = iAHU+2, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Site Location'] )
        iCol += 1
        
        # Building
        wsh.cell( column = iCol, row = iAHU+2, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Building'] )
        iCol += 1
        
        # AHU
        wsh.cell( column = iCol, row = iAHU+2, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Room Height'] )
        iCol += 1
        
        # System Type
        wsh.cell( column = iCol, row = iAHU+2, value = False )
        iCol += 1
        
        # Average Space Height (m)
        wsh.cell( column = iCol, row = iAHU+2, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Room Height'] )
        iCol += 1
        
        # Area (m²)
        wsh.cell( column = iCol, row = iAHU+2, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Room Area'] )
        iCol += 1
        
        # Airflow (m3/s)
        wsh.cell( column = iCol, row = iAHU+2, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Supply Airflow'] )
        iCol += 1
        
        # Current Air Change Rate (Ach)
        wsh.cell( column = iCol, row = iAHU+2, value = wsh.cell(column = iCol, row = 2).value )
        iCol += 1
        
        # Room Occupancy Density
        wsh.cell( column = iCol, row = iAHU+2, value = 0 )
        iCol += 1
        
        # Minimum Fresh Air Percentage
        wsh.cell( column = iCol, row = iAHU+2, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Minimum Fresh Air Percentage'] )
        iCol += 1
        
        # Active Mixing
        wsh.cell( column = iCol, row = iAHU+2, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Active Mixing'] )
        iCol += 1
        
        # Heat Recovery Efficiency (S)
        wsh.cell( column = iCol, row = iAHU+2, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Heat Recovery Efficiency (S)'] )
        iCol += 1
        
        # Heat Recovery Efficiency (L)
        wsh.cell( column = iCol, row = iAHU+2, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Heat Recovery Efficiency (S)'] )
        iCol += 1
        
        # Heating Efficiency
        wsh.cell( column = iCol, row = iAHU+2, value = 1 )
        iCol += 1
        
        # Cooling COP
        wsh.cell( column = iCol, row = iAHU+2, value = 3 )
        iCol += 1
        
        # Supply Fan Total Static Pressure
        wsh.cell( column = iCol, row = iAHU+2, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Supply Fan Total Static Pressure'] )
        iCol += 1
        
        # Return Fan Total Static Pressure
        wsh.cell( column = iCol, row = iAHU+2, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Return Fan Total Static Pressure'] )
        iCol += 1
        
        # Room Cleanliness Class
        wsh.cell( column = iCol, row = iAHU+2, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Room Cleanliness Class'] )
        iCol += 1
        
        # System Type2
        systemType = ''
        for iHVAC in range(1,11):
            if inputs_dicts[str(iAHU)]['model_inputs']['General']['HVAC Module ' + str(iHVAC)] == 0:
                systemType = systemType[0:-2]
                break
            systemType += inputs_dicts[str(iAHU)]['model_inputs']['General']['HVAC Module ' + str(iHVAC)] + '_'

            
        wsh.cell( column = iCol, row = iAHU+2, value = systemType )
        # jump over 5 columns since they are hidden
        iCol += 6
        
        # Fan Power (kWh)
        wsh.cell( column = iCol, row = iAHU+2, value = summary_values_cases[str(iAHU)]['total']['fans_total'] )
        iCol += 1
        
        # Preheat (kWh)
        wsh.cell( column = iCol, row = iAHU+2, value = summary_values_cases[str(iAHU)]['total']['preheat_total'] )
        iCol += 1
        
        # Reheat (kWh)
        reheat = 0
        wsh.cell( column = iCol, row = iAHU+2, value = reheat )
        iCol += 1
        
        # Dehum (kWh)
        wsh.cell( column = iCol, row = iAHU+2, value = summary_values_cases[str(iAHU)]['total']['dehum_total'] )
        iCol += 1
        
        # Hum (kWh)
        wsh.cell( column = iCol, row = iAHU+2, value = summary_values_cases[str(iAHU)]['total']['hum_total'] )
        iCol += 1
        
        # Heat (kWh)
        wsh.cell( column = iCol, row = iAHU+2, value = summary_values_cases[str(iAHU)]['total']['heating_total'] )
        iCol += 1
        
        # Cool (kWh)
        wsh.cell( column = iCol, row = iAHU+2, value = summary_values_cases[str(iAHU)]['total']['cooling_total'] )
        iCol += 1
        
        # Σ Energy (kWh)
        total = summary_values_cases[str(iAHU)]['total']['fans_total'] + summary_values_cases[str(iAHU)]['total']['preheat_total'] + summary_values_cases[str(iAHU)]['total']['dehum_total'] + summary_values_cases[str(iAHU)]['total']['hum_total'] + summary_values_cases[str(iAHU)]['total']['heating_total'] + summary_values_cases[str(iAHU)]['total']['cooling_total'] + reheat
        wsh.cell( column = iCol, row = iAHU+2, value = total )
        iCol += 1
        
        # Volume (m³)
        wsh.cell( column = iCol, row = iAHU+2, value = wsh.cell(column = iCol, row = 2).value )
        iCol += 1
        
        # Space Served
        wsh.cell( column = iCol, row = iAHU+2, value = '' )
        iCol += 1
        
    wb.save(model_file_path)
    
def write_variations_batch_bayer(inputs_dicts, variation_summary_values_cases, model_file_path):
    # pdb.set_trace()
    wb = load_workbook(model_file_path)
    wsh = wb['ECM matrix']
    
    n_AHUs = len(inputs_dicts)
    
    print('clearing old data...')
    # delete old data
    for iRow in range(3,10000):
        for iCol in range(1,56):
            wsh.cell(column = iCol, row = iRow, value = '')
            wsh.cell(column = iCol, row = iRow).border = Border()
    print('finished!')
    iRow = 2
    # write values of variations
    for iAHU in range(n_AHUs):
        nMeasures = len(variation_summary_values_cases[str(iAHU)]['measure']) - 1
        for iMeasure in range(1,nMeasures+1):
            
            # copy existing formulas
            for iCol in range(1,56):
                if not iRow == 2:
                    wsh.cell(column = iCol, row = iRow, value = wsh.cell(column = iCol, row = iRow-1).value)
            iCol = 1
            # pdb.set_trace()
            # Site
            wsh.cell( column = iCol, row = iRow, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Site Location'] )
            iCol += 1
            
            # AHU Ref
            wsh.cell( column = iCol, row = iRow, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['AHU'] )
            iCol += 1
            
            # Run Date
            wsh.cell( column = iCol, row = iRow, value = str(datetime.datetime.now().strftime('%y-%m-%d %H-%M')) )
            iCol += 1
            
            # Room Cleanliness Class
            wsh.cell( column = iCol, row = iRow, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Room Cleanliness Class'] )
            iCol += 1
            
            # Area
            wsh.cell( column = iCol, row = iRow, value = inputs_dicts[str(iAHU)]['model_inputs']['General']['Room Area'] )
            iCol += 1
            
            # Current AHU Configuration
            systemType = ''
            for iHVAC in range(1,11):
                if inputs_dicts[str(iAHU)]['model_inputs']['General']['HVAC Module ' + str(iHVAC)] == 0:
                    systemType = systemType[0:-2]
                    break
                systemType += inputs_dicts[str(iAHU)]['model_inputs']['General']['HVAC Module ' + str(iHVAC)] + '_'
            wsh.cell( column = iCol, row = iRow, value = systemType )
            iCol += 1
            
            # Measure
            wsh.cell( column = iCol, row = iRow, value = variation_summary_values_cases[str(iAHU)]['measure'][iMeasure] )
            iCol += 1
            
            # zeros and ones
            excelMeasures = ['widen_temp_band','widen_humidity_band','reduce_air_flow','incorporate_run_around_coil_heat_recovery','incorporate_mixing_box_recovery','reduce_temp_setpoint','increase_temp_setpoint']
            for iExcelMeasure in excelMeasures:
                wsh.cell( column = iCol, row = iRow, value = int(variation_summary_values_cases[str(iAHU)]['measure_bool'][iMeasure][iExcelMeasure]) )
                iCol += 1
            
            # base scenario and measure scenario
            resultKeys = ['fans', 'preheat', 'dummy', 'dehum', 'hum', 'heating', 'cooling', 'total']
            for iCase in range(2):
                if iCase == 0:
                    iDictValue = 0
                else:
                    iDictValue = iMeasure
                for key in resultKeys:
                    if not key == 'dummy':
                        wsh.cell( column = iCol, row = iRow, value = variation_summary_values_cases[str(iAHU)]['absolute'][key][iDictValue] )
                    else:
                        wsh.cell( column = iCol, row = iRow, value = 0 )
                    iCol += 1
                
                if iCase == 0:
                    iCol += 3
            # pdb.set_trace()
            if iMeasure == nMeasures:
                double = Side(border_style="double")
                for iCell in range(1,56):
                    wsh.cell(row = iRow, column = iCell).border = Border(bottom = double)
                    # True
            iRow += 1
    wb.save(model_file_path)
          
if __name__ == "__main__":  
    
    inputs_path = 'inputs.json'    
    with open(inputs_path, 'r') as infile:
        inputs_dicts = json.load(infile)
    
    #if only one variant
    if not '0' in inputs_dicts.keys():
        inputs_dicts['0'] = inputs_dicts
   
    #if multiple variants
    hvac_obj_cases                  = {}
    detailed_results_cases          = {}
    summary_values_cases            = {}
    variation_summary_values_cases  = {}
    for variant in inputs_dicts:
        hvac_obj_cases[variant]                 = HVAC.HVAC_Internal_Calculation(inputs_dicts[variant])
        detailed_results_cases[variant]         = hvac_obj_cases[variant].get_detailed_results()
        summary_values_cases[variant]           = hvac_obj_cases[variant].get_summary_values()
        print (variant)
        # variation_summary_values_cases[variant] = create_variations(summary_values_cases[variant], inputs_dicts[variant])
        variation_summary_values_cases[variant] = create_variations_batch(summary_values_cases[variant], inputs_dicts[variant])
        # pdb.set_trace()
        
    
    case = 'Master Data.xlsx'
    model_file_path = os.getcwd() + '\\' + case
    
    write_summary_batch(inputs_dicts, summary_values_cases, model_file_path)
    write_variations_batch_bayer(inputs_dicts, variation_summary_values_cases, model_file_path)

    #write_summary_values(model_file_path, 'Model Results Summary', summary_values)        
    #write_detailed_result(model_file_path, 'Model Results Detailed', detailed_results)
    
    # write_variations_result(model_file_path, 'Model Results Variations', variation_summary_values, summary_values)
    
        
        
        
        